import cv2
import numpy as np
import math
import os
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def calculate_t_star(time, rmax):
    """
    t*を計算する関数

    Parameters:
    time (float): 経過時間（s）
    rmax (float): 最大半径（mm）

    Returns:
    float: 計算されたt*の値
    """
    pressure_term = (10**5/1000)**(-0.5)  # 圧力項の計算
    denominator = 0.91468 * rmax * pressure_term * 1000
    if denominator == 0:
        return 0 # ゼロ除算を避ける
    return time / denominator

def find_bubble_points(radius_data):
    """
    気泡の最大半径点と最初の極小点（崩壊点）を見つける関数

    Parameters:
    radius_data (list): 各フレームの半径データ（mm）

    Returns:
    tuple: (最大半径のインデックス, 崩壊点（最初の極小値）のインデックス)
    """
    max_radius = 0
    max_index = -1
    collapse_index = -1

    # まず気泡の発生（0から非0への変化）を見つける
    started = False
    for i, radius in enumerate(radius_data):
        if not started and radius > 0:
            started = True
        if started:
            # 最大半径を更新
            if radius > max_radius:
                max_radius = radius
                max_index = i

    # 最大半径点以降の最初の極小値を見つける
    # この閾値は、崩壊点をより正確に見つけるために調整が必要かもしれません
    radius_threshold = 1.00 
    if max_index != -1:
        for i in range(max_index + 1, len(radius_data)):
            # 現在の値が両隣よりも小さい、かつ閾値以下であれば極小点と判断
            if i > 0 and i < len(radius_data) - 1:
                if radius_data[i] < radius_threshold and \
                   radius_data[i] < radius_data[i-1] and \
                   radius_data[i] <= radius_data[i+1]:
                    collapse_index = i
                    break
            # 最後の要素の場合
            elif i == len(radius_data) - 1: 
                if radius_data[i] < radius_data[i-1] and \
                   radius_data[i] < radius_threshold:
                    collapse_index = i
                    break

    # もし極小値が見つからず、最後に0になる点があればそれを崩壊点とする
    if collapse_index == -1:
        for i in range(max_index + 1, len(radius_data)):
            if radius_data[i] == 0:
                collapse_index = i
                break

    return max_index, collapse_index

def calculate_bubble_properties(image_path, output_path, calibration, wall_x_pixel=None):
    """
    単一の画像から気泡の特性を計算する関数

    Parameters:
    image_path (str): 入力画像ファイルパス
    output_path (str): 結果画像を保存するパス
    calibration (float): ピクセルからミリメートルへの変換係数 (pix/mm)
    wall_x_pixel (int, optional): 壁の位置を示すX座標（ピクセル）。指定すると画像に線を描画。

    Returns:
    dict or None: 気泡の体積、半径、重心、境界座標（mm単位）。検出できなかった場合はNone。
    """
    if not os.path.isfile(image_path):
        print(f"画像ファイルが存在しません: {image_path}")
        return None

    try:
        # 日本語パス対応のためにcv2.imdecodeを使用
        img = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)

        if img is None:
            print(f"画像の読み込みに失敗しました: {image_path}")
            return None
    except Exception as e:
        print(f"画像読み込み時にエラーが発生しました: {image_path} - {str(e)}")
        return None

    # 画像処理
    blurred = cv2.GaussianBlur(img, (37, 37), 0)
    binary = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 31, 2)
    kernel = np.ones((3,3), np.uint8)
    binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=2)

    # 輪郭検出
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    img_result = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

    # 壁の描画（指定がある場合）
    if wall_x_pixel is not None:
        for y_coord in range(0, img.shape[0], 10):
            cv2.line(img_result, (int(wall_x_pixel), y_coord), (int(wall_x_pixel), min(y_coord + 5, img.shape[0])), (0, 0, 255), 2)

    if contours:
        # 最大の輪郭を気泡として選択
        bubble = max(contours, key=cv2.contourArea)
        x, y, w, h = cv2.boundingRect(bubble)

        # マスクを作成し、輪郭内部を塗りつぶす
        mask = np.zeros(binary.shape, dtype=np.uint8)
        cv2.drawContours(mask, [bubble], 0, 255, -1)

        # 気泡の体積と重心を計算
        volume = 0 # ピクセル単位の体積
        total_cx, total_cy = 0, 0
        total_area = 0

        for i in range(h):
            slice_mask = mask[y+i, x:x+w]
            if np.sum(slice_mask) > 0:
                non_zero = np.nonzero(slice_mask)[0]
                left = non_zero[0]
                right = non_zero[-1]

                cx = x + left + (right - left) // 2
                cy = y + i

                radius_pixel = (right - left) / 2
                slice_volume = math.pi * radius_pixel**2 # ピクセル単位の微小円盤の体積（2D）

                volume += slice_volume

                slice_area = np.sum(slice_mask > 0)
                total_cx += cx * slice_area
                total_cy += cy * slice_area
                total_area += slice_area

        if total_area > 0:
            center_of_mass = (int(total_cx / total_area), int(total_cy / total_area))
        else:
            center_of_mass = (x + w // 2, y + h // 2)

        # ピクセル単位の値をミリメートル単位に変換
        volume_mm3 = volume / (calibration**3)
        equivalent_radius_mm = (3 * volume_mm3 / (4 * math.pi))**(1/3) if volume_mm3 > 0 else 0
        x_min_mm = x / calibration
        x_max_mm = (x+w) / calibration
        y_min_mm = y / calibration
        y_max_mm = (y+h) / calibration

        # 小さすぎる体積の気泡は無視（必要に応じて閾値設定）
        min_volume = 0 
        if volume_mm3 < min_volume:
            return {
                'volume': 0,
                'radius': 0,
                'center': (center_of_mass[0]/calibration, center_of_mass[1]/calibration),
                'x_min_mm': 0, 'x_max_mm': 0, 'y_min_mm': 0, 'y_max_mm': 0
            }

        # 結果画像に輪郭と重心を描画
        cv2.drawContours(img_result, [bubble], 0, (0, 255, 0), 2)
        cv2.circle(img_result, center_of_mass, 5, (0, 0, 255), -1)

        # 結果画像を保存
        try:
            _, buf = cv2.imencode('.bmp', img_result)
            buf.tofile(output_path) # 日本語パス対応
        except Exception as e:
            print(f"結果の保存に失敗しました: {output_path} - {str(e)}")

        return {
            'volume': volume_mm3,
            'radius': equivalent_radius_mm,
            'center': (center_of_mass[0]/calibration, center_of_mass[1]/calibration),
            'x_min_mm': x_min_mm,
            'x_max_mm': x_max_mm,
            'y_min_mm': y_min_mm,
            'y_max_mm': y_max_mm
        }
    else:
        # 気泡が検出されなかった場合
        try:
            _, buf = cv2.imencode('.bmp', img_result)
            buf.tofile(output_path) # 日本語パス対応
        except Exception as e:
            print(f"結果の保存に失敗しました: {output_path} - {str(e)}")
        print(f"気泡を検出できませんでした: {image_path}")
        return None

def update_excel_for_folders(folders_to_update, base_path, calibration, time_interval, start_image_num, excel_file_name='analysis_result.xlsx'):
    """
    指定されたフォルダの気泡データを解析し、既存のExcelファイルを更新する関数。

    Parameters:
    folders_to_update (list): 更新対象のフォルダ番号のリスト
    base_path (str): 解析する画像のルートディレクトリ
    calibration (float): ピクセルからミリメートルへの変換係数 (pix/mm)
    time_interval (float): 各フレーム間の時間間隔（s）
    start_image_num (int): t*=0とする画像番号（1始まり）
    excel_file_name (str): 更新するExcelファイルの名前
    """
    # 基準値ファイルのパスを正規化
    reference_path = os.path.join(base_path, 'reference.xlsx').replace('/', '\\')

    if not os.path.isfile(reference_path):
        print(f"基準値ファイルが存在しません: {reference_path}")
        return

    try:
        reference_df = pd.read_excel(reference_path)
        reference_dict = dict(zip(reference_df['Folder'], reference_df['base_x(pix)']))
        print("基準値の読み込みが完了しました。")
    except Exception as e:
        print(f"基準値ファイルの読み込みに失敗しました: {str(e)}")
        return

    results_base_path = os.path.join(base_path, "results").replace('/', '\\')
    os.makedirs(results_base_path, exist_ok=True)

    excel_path = os.path.join(results_base_path, excel_file_name).replace('/', '\\')

    # 既存のExcelファイルを読み込む
    try:
        wb = load_workbook(excel_path)
        ws = wb.active # アクティブなシートを選択
        print(f"既存のExcelファイル {excel_path} を読み込みました。")
    except FileNotFoundError:
        print(f"エラー: Excelファイル {excel_path} が見つかりません。")
        print("最初にすべてのフォルダを処理して初期のExcelファイルを生成しておく必要があります。")
        return
    except Exception as e:
        print(f"既存のExcelファイルの読み込みに失敗しました: {str(e)}")
        return

    # ヘッダー行（2行目）から各フォルダの開始列を特定する
    # このマップは、各フォルダのデータブロックの 't*' 列のExcel列番号（1始まり）を格納します。
    folder_t_star_col_map = {} 
    max_excel_col = ws.max_column
    for col_idx in range(1, max_excel_col + 1):
        # Excelの2行目のセル値をチェック（Folder番号が格納されているセル）
        cell_value = ws.cell(row=2, column=col_idx).value 
        if isinstance(cell_value, (int, float)):
            # Folder番号のセル (col_idx) から 't*' の列までの正しいオフセットを設定
            # 過去の試行錯誤とユーザーからのフィードバックに基づき、-2 が正しいオフセットであると判断
            folder_t_star_col_map[int(cell_value)] = col_idx - 2 
            
    # print(f"特定されたフォルダ列マッピング (t*の列番号): {folder_t_star_col_map}") # デバッグ出力削除

    # 各フォルダを処理
    for folder_num in folders_to_update:
        folder_name = str(folder_num)
        folder_path = os.path.join(base_path, folder_name).replace('/', '\\')

        if not os.path.exists(folder_path):
            print(f"フォルダが存在しません: {folder_path}")
            continue

        result_folder = os.path.join(results_base_path, folder_name).replace('/', '\\')
        os.makedirs(result_folder, exist_ok=True)

        print(f"--- フォルダ {folder_name} の処理を開始・更新中 ---")

        # まず半径データを収集（t*計算のために最大半径が必要なため）
        radius_data = []
        current_wall_x_pixel = reference_dict.get(folder_num, None)

        # ファイルリストをソートして処理順を保証
        image_files = sorted([f for f in os.listdir(folder_path) if f.endswith('.bmp')])

        # 全画像から半径データを一度収集（gamma計算用の一時処理）
        for file_name in image_files:
            image_path = os.path.join(folder_path, file_name).replace('/', '\\')
            # 結果画像を生成する必要がないため、ダミーのoutput_pathを使用
            dummy_output_path_for_radius = os.path.join(result_folder, "temp_rad_calc.bmp").replace('/', '\\')
            result = calculate_bubble_properties(image_path, dummy_output_path_for_radius, calibration, wall_x_pixel=current_wall_x_pixel)
            radius_data.append(result['radius'] if result else 0)
            # 不要な一時ファイルを削除
            if os.path.exists(dummy_output_path_for_radius):
                os.remove(dummy_output_path_for_radius)

        if not radius_data:
            print(f"フォルダ {folder_name} にデータがありません")
            continue

        # 最大半径点と崩壊点を特定
        max_index, collapse_index = find_bubble_points(radius_data)

        rmax_for_t_star = 0
        if max_index != -1 and max_index < len(radius_data):
            rmax_for_t_star = radius_data[max_index]

        # Excelの更新対象列を特定
        if folder_num not in folder_t_star_col_map:
            print(f"警告: フォルダ {folder_num} のデータがExcelファイルに見つかりません。スキップします。")
            continue

        # excel_col_t_star は、現在のフォルダの 't*' データが始まるExcelの列番号（1始まり）
        excel_col_t_star = folder_t_star_col_map[folder_num] 
        # print(f"  --> このフォルダの 't*' データ開始列 (Excel 1-based): {excel_col_t_star}") # デバッグ出力削除


        # フォルダ情報（Rmax, γ）を更新
        # Rmaxとγは、「半径の列」（'t*' の列から2つ右）に設定
        col_for_rmax_gamma = excel_col_t_star + 2
        ws.cell(row=3, column=col_for_rmax_gamma).value = rmax_for_t_star 
        # print(f"  --> Rmax/Gamma 書き込み列 (Excel 1-based): {col_for_rmax_gamma}") # デバッグ出力削除
        
        # γの計算と更新
        gamma = 0
        if len(radius_data) >= 8 and folder_num in reference_dict:
            base_x = reference_dict[folder_num]
            # 11番目のデータ（インデックス10）が存在するか確認
            if len(radius_data) > 10: 
                # γ計算のため、改めて11番目の画像の重心X座標を計算する
                initial_x_image_path = os.path.join(folder_path, image_files[10]).replace('/', '\\')
                dummy_output_path_gamma = os.path.join(result_folder, "temp_gamma_calc.bmp").replace('/', '\\')
                initial_image_result = calculate_bubble_properties(initial_x_image_path, dummy_output_path_gamma, calibration, wall_x_pixel=current_wall_x_pixel)
                
                if initial_image_result and rmax_for_t_star > 0:
                    initial_x_mm = initial_image_result['center'][0] # mm単位の重心X
                    gamma = (initial_x_mm - base_x / calibration) / rmax_for_t_star
                    # 不要な一時ファイルを削除
                    if os.path.exists(dummy_output_path_gamma):
                        os.remove(dummy_output_path_gamma)
                
        ws.cell(row=4, column=col_for_rmax_gamma).value = gamma 


        # データ行を更新
        for idx, file_name in enumerate(image_files):
            if idx >= 120: # 120フレームまで
                break

            image_path = os.path.join(folder_path, file_name).replace('/', '\\')
            output_path = os.path.join(result_folder, f"{os.path.splitext(file_name)[0]}_result.bmp").replace('/', '\\')

            # ここで画像解析を再度実行し、最新の結果を取得
            result = calculate_bubble_properties(image_path, output_path, calibration, wall_x_pixel=current_wall_x_pixel)

            data_row_excel = idx + 6 # Excelのデータ開始行番号（1始まり）

            t_star = 0
            # t*=0とする画像番号以降でt*を計算
            if result and idx >= start_image_num - 1 and rmax_for_t_star != 0:
                elapsed_time = (idx - (start_image_num - 1)) * time_interval
                t_star = calculate_t_star(elapsed_time, rmax_for_t_star)

            # 各データ列を更新
            # ws.cell(row=行番号, column=列番号).value = 値
            # 列番号は Excel の1始まりの番号
            # t* は excel_col_t_star
            # 体積は t* から +1
            # 半径は t* から +2
            # 重心Xは t* から +3
            # 重心Yは t* から +4
            if result:
                ws.cell(row=data_row_excel, column=excel_col_t_star).value = t_star
                ws.cell(row=data_row_excel, column=excel_col_t_star + 1).value = result['volume']
                ws.cell(row=data_row_excel, column=excel_col_t_star + 2).value = result['radius']
                ws.cell(row=data_row_excel, column=excel_col_t_star + 3).value = result['center'][0]
                ws.cell(row=data_row_excel, column=excel_col_t_star + 4).value = result['center'][1]
                ws.cell(row=data_row_excel, column=excel_col_t_star + 5).value = result['x_min_mm']
                ws.cell(row=data_row_excel, column=excel_col_t_star + 6).value = result['x_max_mm']
                ws.cell(row=data_row_excel, column=excel_col_t_star + 7).value = result['y_min_mm']
                ws.cell(row=data_row_excel, column=excel_col_t_star + 8).value = result['y_max_mm']
            else:
                # 気泡が検出されない場合は0または空欄にする
                ws.cell(row=data_row_excel, column=excel_col_t_star).value = t_star
                ws.cell(row=data_row_excel, column=excel_col_t_star + 1).value = 0
                ws.cell(row=data_row_excel, column=excel_col_t_star + 2).value = 0
                ws.cell(row=data_row_excel, column=excel_col_t_star + 3).value = 0
                ws.cell(row=data_row_excel, column=excel_col_t_star + 4).value = 0
                ws.cell(row=data_row_excel, column=excel_col_t_star + 5).value = 0
                ws.cell(row=data_row_excel, column=excel_col_t_star + 6).value = 0
                ws.cell(row=data_row_excel, column=excel_col_t_star + 7).value = 0
                ws.cell(row=data_row_excel, column=excel_col_t_star + 8).value = 0


            # 最大半径点と崩壊点のセルに色付け
            # 体積の列のみに色付けを適用する
            target_col_excel_for_color = excel_col_t_star + 1 # 体積の列 (t*から+1)
            
            # ターゲットセルを取得
            cell_to_color = ws.cell(row=data_row_excel, column=target_col_excel_for_color)

            # このセルの既存の塗りつぶしを確実にクリアする (重要！)
            cell_to_color.fill = PatternFill(fill_type=None)

            # 条件に応じて新しい色を適用
            if idx == max_index:
                cell_to_color.fill = PatternFill(fgColor='FF0000', fill_type='solid') # 赤色
                # print(f"  Frame {idx}: 最大半径点検出。列 {target_col_excel_for_color} (体積) を赤色に設定。") # デバッグ出力削除
            elif idx == collapse_index:
                cell_to_color.fill = PatternFill(fgColor='FFFF00', fill_type='solid') # 黄色
                # print(f"  Frame {idx}: 崩壊点検出。列 {target_col_excel_for_color} (体積) を黄色に設定。") # デバッグ出力削除


    # Excelファイルを保存
    try:
        wb.save(excel_path)
        print(f"Excelファイル {excel_path} が正常に更新されました。")
    except Exception as e:
        print(f"Excelファイルの保存に失敗しました: {str(e)}")

    print("選択されたフォルダの処理とExcel更新が完了しました。")


if __name__ == "__main__":
    # 処理の設定
    base_path = r'C:\Research\exp_data\20250611' # raw文字列として指定
    calibration = 39.4
    time_interval = 10  # 時間間隔s（秒）
    start_image_num = 7 # t*=0とする画像番号（1始まり）

    # 更新したいフォルダ番号のリストを指定
    folders_to_reanalyze = [63] # 例としてフォルダ2, 3を更新

    try:
        update_excel_for_folders(folders_to_reanalyze, base_path, calibration, time_interval, start_image_num,
                                 excel_file_name="9_analysis_result.xlsx")
    except Exception as e:
        print(f"プログラム実行中にエラーが発生しました: {str(e)}")