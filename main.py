import cv2
import numpy as np
import math
import os
import pandas as pd
from openpyxl.styles import PatternFill

def calculate_t_star(time, rmax):
    """
    t*を計算する関数
    time: 経過時間（s）
    rmax: 最大半径（mm）
    """
    rho = 1000  # 水の密度 (kg/m³)
    delta_p = 1e5  # 圧力差 (Pa)
    
    denominator = 0.91468 * (rmax / 1000) * (rho / delta_p)**0.5
    return time / denominator

def find_bubble_points(radius_data):
    """
    気泡の最大半径点と最初の極小点（崩壊点）を見つける
    """
    max_radius = 0
    max_index = -1
    collapse_index = -1

    # まず気泡の発生（0から非0への変化）を見つける
    started = False
    start_bubble_index = -1
    for i, radius in enumerate(radius_data):
        if not started and radius > 0:
            started = True
            start_bubble_index = i
        if started:
            # 最大半径を更新
            if radius > max_radius:
                max_radius = radius
                max_index = i

    # 最大半径点以降の最初の極小値を見つける
    # ただし、極小値を探す範囲を最大半径点以降に限定
    radius_threshold = 1.00
    if max_index != -1:
        # 最大半径点からデータ終了までを探索
        for i in range(max_index + 1, len(radius_data)):
            # 谷底を探す条件: 現在の値が両隣よりも小さい
            # ただし、配列の境界チェックも必要
            if i > 0 and i < len(radius_data) - 1:
                if radius_data[i] < radius_threshold and radius_data[i] < radius_data[i-1] and radius_data[i] <= radius_data[i+1]:
                    collapse_index = i
                    break # 最初の極小値を見つけたらループを抜ける
            elif i == len(radius_data) - 1: # 最後の要素の場合
                if radius_data[i] < radius_data[i-1]:
                    collapse_index = i
                    break
    
    # もし極小値が見つからず、最後に0になる点があればそれを崩壊点とする（元のロジックの補完）
    if collapse_index == -1:
        for i in range(max_index + 1, len(radius_data)):
            if radius_data[i] == 0:
                collapse_index = i
                break
    
    return max_index, collapse_index

# wall_x_pixel を引数に追加
def calculate_bubble_properties(image_path, output_path, calibration, blur_x, blur_y, binary_s, wall_x_pixel=None):
    # ファイルの存在確認
    if not os.path.isfile(image_path):
        print(f"画像ファイルが存在しません: {image_path}")
        return
    
    # 画像読み込み - ファイルパスを適切に処理
    try:
        # Windows環境でのファイルパス問題に対処
        # バックスラッシュを含むパスを raw 文字列としてエンコード
        img = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
        
        if img is None:
            print(f"画像の読み込みに失敗しました: {image_path}")
            return
    except Exception as e:
        print(f"画像読み込み時にエラーが発生しました: {image_path} - {str(e)}")
        return

    # 画像の前処理
    blurred = cv2.GaussianBlur(img, (blur_x, blur_y), 0)
    
    # 適応的閾値処理
    binary = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, binary_s, 2)

    # モルフォロジー演算
    kernel = np.ones((3,3), np.uint8)
    binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=2)

    # 輪郭検出
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # 結果の可視化用の画像を用意
    img_result = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

    # 壁面位置の描画 (wall_x_pixelが指定されている場合のみ)
    if wall_x_pixel is not None:
        # 点線を描画
        # OpenCVのlineTypeには点線タイプが直接ないので、手動で描画
        for y_coord in range(0, img.shape[0], 10): # 10ピクセル描画、10ピクセル空白の例
            cv2.line(img_result, (int(wall_x_pixel), y_coord), (int(wall_x_pixel), min(y_coord + 5, img.shape[0])), (0, 0, 255), 2) # 赤色、3ピクセル太さ

    if contours:
        # 最大の輪郭を気泡として扱う
        bubble = max(contours, key=cv2.contourArea)

        # バウンディングボックスを取得(左上の座標をx,y、横と縦のサイズをw,h)
        x, y, w, h = cv2.boundingRect(bubble)

        # 気泡のマスクを作成
        mask = np.zeros(binary.shape, dtype=np.uint8)
        cv2.drawContours(mask, [bubble], 0, 255, -1)

        # === X軸に沿った積分（垂直円盤スタック近似）のための初期化 ===
        volume_agarose_sum = 0   # Vagarose (左側, pixel^3 相当)
        volume_water_sum = 0     # Vwater (右側, pixel^3 相当)
        
        # 重心計算用（分子: モーメント = x * Vslice または y_center * Vslice）
        moment_x_agarose = 0     
        moment_y_agarose = 0
        moment_x_water = 0
        moment_y_water = 0
        
        # 全体の重心計算のための累積 (面積ベース)
        total_cx, total_cy = 0, 0
        total_area_overall = 0
        
        wall_x = int(wall_x_pixel) if wall_x_pixel is not None else -1 

        # X軸に沿ってスライス（垂直の列を走査）
        for j in range(w):
            x_abs = x + j # 絶対X座標
            
            # スライス（垂直な列）を取得
            slice_mask = mask[y:y+h, x_abs] 
            
            if np.sum(slice_mask) > 0:
                # スライスの輪郭（上端と下端）を見つける
                non_zero_y = np.nonzero(slice_mask)[0]
                y_top_rel = non_zero_y[0]   # yを基準とした相対座標
                y_bottom_rel = non_zero_y[-1]
                
                # 半径 r(x) と中心 y_center(x) を計算 (ピクセル単位)
                slice_height = y_bottom_rel - y_top_rel + 1
                r_pixel = slice_height / 2
                y_center_abs = y + y_top_rel + r_pixel - 0.5 # 絶対Y座標の中心（ピクセル）

                # 体積素片 Vslice(x) を計算 (π * r^2 * Δx, Δx=1 pixel)
                v_slice_pixel3 = math.pi * r_pixel**2 * 1 

                # 全体の重心計算（面積ベース）
                slice_area = slice_height * 1
                total_cx += x_abs * slice_area
                total_cy += y_center_abs * slice_area
                total_area_overall += slice_area

                # === Vagarose / Vwater の分割とモーメントの計算 ===
                if wall_x != -1:
                    # Vagarose (左側, x < wall_x)
                    if x_abs < wall_x: 
                        volume_agarose_sum += v_slice_pixel3
                        moment_x_agarose += x_abs * v_slice_pixel3
                        moment_y_agarose += y_center_abs * v_slice_pixel3
                        
                    # Vwater (右側, x >= wall_x)
                    elif x_abs >= wall_x: 
                        volume_water_sum += v_slice_pixel3
                        moment_x_water += x_abs * v_slice_pixel3
                        moment_y_water += y_center_abs * v_slice_pixel3
                else:
                    # wall_xがない場合、全体の体積として累積
                    volume_agarose_sum += v_slice_pixel3 
                    
        # 気泡全体の重心を計算 (面積ベース)
        if total_area_overall > 0:
            center_of_mass = (int(total_cx / total_area_overall), int(total_cy / total_area_overall))
        else:
            center_of_mass = (x + w // 2, y + h // 2)

        # === キャリブレーションを最後にまとめて適用 ===
        
        # 全体の体積を Vagarose + Vwater で定義
        volume_total_pixel3 = volume_agarose_sum + volume_water_sum
        
        # 1. 体積 (mm^3)
        volume_mm3 = volume_total_pixel3 / (calibration**3)
        volume_agarose_mm3 = volume_agarose_sum / (calibration**3)
        volume_water_mm3 = volume_water_sum / (calibration**3)
        
        # 新しい体積に基づく等価半径
        equivalent_radius_mm = (3 * volume_mm3 / (4 * math.pi))**(1/3) 
        
        # 【追加】アスペクト比の計算 (h/w)
        aspect_ratio = h / w if w != 0 else 0
        
        x_min_mm = x / calibration
        x_max_mm = (x+w) / calibration
        y_min_mm = y / calibration
        y_max_mm = (y+h) / calibration
        
        # 2. 重心 (mm)
        # Zagarose (X)
        center_x_agarose_mm = (moment_x_agarose / volume_agarose_sum) / calibration if volume_agarose_sum > 0 else 0
        # Yagarose (Y)
        center_y_agarose_mm = (moment_y_agarose / volume_agarose_sum) / calibration if volume_agarose_sum > 0 else 0
        
        # Zwater (X)
        center_x_water_mm = (moment_x_water / volume_water_sum) / calibration if volume_water_sum > 0 else 0
        # Ywater (Y)
        center_y_water_mm = (moment_y_water / volume_water_sum) / calibration if volume_water_sum > 0 else 0
        
         # 結果の可視化 (気泡の外形と重心)
        cv2.drawContours(img_result, [bubble], 0, (0, 255, 0), 2) # 緑色
        cv2.circle(img_result, center_of_mass, 5, (0, 0, 255), -1) # 赤色

        # 結果を保存
        try:
            _, buf = cv2.imencode('.bmp', img_result)
            buf.tofile(output_path)
        except Exception as e:
            print(f"結果の保存に失敗しました: {output_path} - {str(e)}")
            
        return {
            'volume': volume_mm3, 
            'radius': equivalent_radius_mm, 
            'center': (center_of_mass[0]/calibration, center_of_mass[1]/calibration),
            'x_min_mm': x_min_mm,
            'x_max_mm':x_max_mm,
            'y_min_mm':y_min_mm,
            'y_max_mm':y_max_mm,
            # Vagarose, Vwater
            'volume_agarose': volume_agarose_mm3,
            'volume_water': volume_water_mm3,
            # Zagarose (X), Yagarose (Y)
            'center_x_agarose': center_x_agarose_mm,
            'center_y_agarose': center_y_agarose_mm,
            # Zwater (X), Ywater (Y)
            'center_x_water': center_x_water_mm,
            'center_y_water': center_y_water_mm,
            # 【追加】アスペクト比
            'aspect_ratio': aspect_ratio
        }
    else:
        # 気泡が検出されない場合でも、壁面が描画された画像を保存
        try:
            _, buf = cv2.imencode('.bmp', img_result)
            buf.tofile(output_path)
        except Exception as e:
            print(f"結果の保存に失敗しました: {output_path} - {str(e)}")
        print(f"気泡を検出できませんでした: {image_path}")
        return None

def process_all_folders(start_folder, end_folder, base_path, calibration, time_interval, start_image_num, blur_x, blur_y, binary_s, excel_file_name='analysis_result.xlsx'):
    # 基準値ファイルのパスを正規化 - バックスラッシュを使用
    reference_path = os.path.join(base_path, 'reference.xlsx').replace('/', '\\')
    
    # 基準値ファイルの存在確認
    if not os.path.isfile(reference_path):
        print(f"基準値ファイルが存在しません: {reference_path}")
        return
        
    # 基準値の読み込み
    try:
        reference_df = pd.read_excel(reference_path)
        # reference_dictには、フォルダ番号をキーとして'base_x(pix)'の値を格納
        reference_dict = dict(zip(reference_df['Folder'], reference_df['base_x(pix)']))
        print("基準値の読み込みが完了しました")
    except Exception as e:
        print(f"基準値ファイルの読み込みに失敗しました: {str(e)}")
        return

    # 結果保存用のフォルダを作成 - バックスラッシュを使用
    results_base_path = os.path.join(base_path, "results").replace('/', '\\')
    os.makedirs(results_base_path, exist_ok=True)
    
    # Excelファイルの作成準備
    excel_path = os.path.join(results_base_path, excel_file_name).replace('/', '\\')
    
    # データを格納するための辞書
    all_data = {}
    max_radii = {}    # 各フォルダの最大半径を保存
    gammas = {}       # 各フォルダのγ値を保存

    # 各フォルダを処理
    for folder_num in range(start_folder, end_folder + 1):
        folder_name = str(folder_num)
        folder_path = os.path.join(base_path, folder_name).replace('/', '\\')
        
        if not os.path.exists(folder_path):
            print(f"フォルダが存在しません: {folder_path}")
            continue
            
        # 結果保存用のフォルダを作成
        result_folder = os.path.join(results_base_path, folder_name).replace('/', '\\')
        os.makedirs(result_folder, exist_ok=True)
        
        print(f"フォルダ {folder_name} の処理を開始...")
        
        folder_data = []
        max_radius = 0

        # まず半径データを収集（t*計算のために最大半径が必要なため）
        radius_data = []
        # 現在のフォルダの壁面X座標を取得
        current_wall_x_pixel = reference_dict.get(folder_num, None)

        for file_name in sorted(os.listdir(folder_path)):
            if file_name.endswith('.bmp'):
                image_path = os.path.join(folder_path, file_name).replace('/', '\\')
                output_path = os.path.join(result_folder, f"{os.path.splitext(file_name)[0]}_result.bmp").replace('/', '\\')
                
                try:
                    # ここでは壁面描画のためだけに calculate_bubble_properties を呼び出す (結果は一旦無視)
                    # 壁面情報も渡すように変更
                    result = calculate_bubble_properties(image_path, output_path, calibration, blur_x, blur_y, binary_s, wall_x_pixel=current_wall_x_pixel)
                    if result:
                        radius_data.append(result['radius'])
                    else:
                        radius_data.append(0) # 結果が取得できない場合は0を追加
                except Exception as e:
                    print(f"エラーが発生しました - {file_name}: {str(e)}")
                    radius_data.append(0) # エラー時も0を追加してインデックスの整合性を保つ

        # データがない場合は次のフォルダへ
        if not radius_data:
            print(f"フォルダ {folder_name} にデータがありません")
            continue
            
        # 最大半径と崩壊点のインデックスを取得
        max_index, collapse_index = find_bubble_points(radius_data)

        # t*計算のための最大半径を取得 (0でなければ)
        rmax_for_t_star = 0
        if max_index != -1 and max_index < len(radius_data):
            rmax_for_t_star = radius_data[max_index]

        # 実際のデータ収集とExcelデータ生成のためのループ
        for idx, file_name in enumerate(sorted(os.listdir(folder_path))):
            if file_name.endswith('.bmp'):
                image_path = os.path.join(folder_path, file_name).replace('/', '\\')
                output_path = os.path.join(result_folder, f"{os.path.splitext(file_name)[0]}_result.bmp").replace('/', '\\')
                
                try:
                    result = calculate_bubble_properties(image_path, output_path, calibration, blur_x, blur_y, binary_s, wall_x_pixel=current_wall_x_pixel)
                    if result:
                        # t*の計算
                        if idx < start_image_num - 1 or rmax_for_t_star == 0:
                            t_star = 0
                        else:
                            elapsed_time = (idx - (start_image_num - 1)) * time_interval
                            t_star = calculate_t_star(elapsed_time, rmax_for_t_star)
                        
                        folder_data.append({
                            't_star': t_star,
                            'volume': result['volume'],
                            'radius': result['radius'],
                            'center_x': result['center'][0],
                            'center_y': result['center'][1],
                            'x_min_mm':result['x_min_mm'],
                            'x_max_mm':result['x_max_mm'],
                            'y_min_mm':result['y_min_mm'],
                            'y_max_mm':result['y_max_mm'],
                            # Vagarose, Vwater
                            'volume_agarose': result['volume_agarose'],
                            'volume_water': result['volume_water'],
                            # Zagarose (X), Yagarose (Y)
                            'center_x_agarose': result['center_x_agarose'],
                            'center_y_agarose': result['center_y_agarose'],
                            # Zwater (X), Ywater (Y)
                            'center_x_water': result['center_x_water'],
                            'center_y_water': result['center_y_water'],
                            # 【追加】アスペクト比
                            'aspect_ratio': result['aspect_ratio']
                        })
                        max_radius = max(max_radius, result['radius'])
                except Exception as e:
                    print(f"エラーが発生しました - {file_name}: {str(e)}")
                    # エラー時はデータを追加しない (スキップ)
        
        # データが少なすぎる場合は次のフォルダへ
        if len(folder_data) < 5:
            print(f"フォルダ {folder_name} の有効なデータが不足しています")
            continue
            
        # 120フレームまでのデータを保存
        all_data[folder_name] = folder_data[:120]
        max_radii[folder_name] = max_radius

        # γの計算
        if len(folder_data) >= 8 and folder_num in reference_dict:
            base_x = reference_dict[folder_num]
            # folder_dataのインデックスが存在するか確認
            if len(folder_data) > 10: # 11番目のデータ (インデックス10) が存在するか確認
                initial_x = folder_data[10]['center_x']
                # 分母が0になるのを防止
                if max_radius > 0:
                    gamma = (initial_x - base_x/calibration) / max_radius
                else:
                    gamma = 0
                gammas[folder_name] = gamma
            else:
                gammas[folder_name] = 0 # データ不足の場合は0とする


    # 処理したフォルダがない場合
    if not all_data:
        print("処理できるデータがありませんでした")
        return
        
    # Excelファイルのデータ作成
    num_folders = len(all_data)
    num_cols = 1 + num_folders * 16
    # ヘッダー行と情報行の分も考慮して行数を調整
    df = pd.DataFrame(np.nan, index=range(120 + 5), columns=range(num_cols)) # データ120行 + ヘッダー5行

    # A列に連番 (0-119) を追加
    df.iloc[5:125, 0] = [int(i) for i in range(120)]

    # 最初の列にFolder, Rmax, γを設定
    df.iloc[1, 0] = 'Folder'
    df.iloc[2, 0] = 'Rmax'
    df.iloc[3, 0] = 'γ'

     # 各フォルダのデータを配置
    folder_idx = 0
    for folder_num in range(start_folder, end_folder + 1):
        folder_name = str(folder_num)
        if folder_name not in all_data:
            continue
            
        col_start = 1 + folder_idx * 16 # 各フォルダの開始列を 15 に変更

        df.iloc[1, col_start + 2] = folder_num
        df.iloc[2, col_start + 2] = max_radii.get(folder_name, 0)
        df.iloc[3, col_start + 2] = gammas.get(folder_name, 0)

        # カラム名を16個に変更
        df.iloc[4, col_start:col_start+16] = [
            't*', '体積', '半径', '重心X', '重心Y', 
            'x最小値', 'x最大値', 'y最小値', 'y最大値', 
            'Vagarose', 'Vwater', 'Xagarose', 'Xwater', 'Yagarose', 
            'Ywater', 'AR' # ARを追加
        ]
        # (Vx_min/maxは未使用だがカラム数を保持)

        # データを配置（5行目から）
        for row_idx, data in enumerate(all_data[folder_name]):
            if row_idx < 120: 
                data_row = row_idx + 5
                df.iloc[data_row, col_start] = data['t_star']
                df.iloc[data_row, col_start+1] = data['volume']
                df.iloc[data_row, col_start+2] = data['radius']
                df.iloc[data_row, col_start+3] = data['center_x']
                df.iloc[data_row, col_start+4] = data['center_y']
                df.iloc[data_row, col_start+5] = data['x_min_mm']
                df.iloc[data_row, col_start+6] = data['x_max_mm']
                df.iloc[data_row, col_start+7] = data['y_min_mm']
                df.iloc[data_row, col_start+8] = data['y_max_mm']
                # Vagarose, Vwater
                df.iloc[data_row, col_start+9] = data['volume_agarose']
                df.iloc[data_row, col_start+10] = data['volume_water']
                # Zagarose (X), Xwater(X)
                df.iloc[data_row, col_start+11] = data['center_x_agarose']
                df.iloc[data_row, col_start+12] = data['center_x_water']
                # Yagarose(Y), Ywater (Y)
                df.iloc[data_row, col_start+13] = data['center_y_agarose']
                df.iloc[data_row, col_start+14] = data['center_y_water']
                #アスペクト比
                df.iloc[data_row, col_start+15] = data['aspect_ratio'] 

        folder_idx += 1

    # Excelファイルに保存
    try:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        df.to_excel(writer, index=False, header=False)
        
        # ワークシートを取得
        worksheet = writer.sheets['Sheet1']
        
        # 各フォルダの列に対して色付け
        color_folder_idx = 0
        for folder_num in range(start_folder, end_folder + 1):
            folder_name = str(folder_num)
            if folder_name not in all_data:
                continue
                
            col_start = 1 + color_folder_idx * 16  # 各フォルダの開始列
            radius_data = [data['radius'] for data in all_data[folder_name]]
            max_index, collapse_index = find_bubble_points(radius_data)
            
            # Excelの行番号は1から始まるため +1
            # DataFrameのインデックスは0から始まるため、データ開始行のインデックス5を考慮
            
            # インデックスが範囲内かチェック
            if 0 <= max_index < len(radius_data):
                # 最大半径時のセルを赤色で塗る
                max_row_excel = max_index + 6  # Excelの行番号
                cell = worksheet.cell(row=max_row_excel, column=col_start+2)  # 半径の列
                cell.fill = PatternFill(fgColor='FF0000', fill_type='solid')
            
            # インデックスが範囲内かチェック
            if 0 <= collapse_index < len(radius_data):
                # 崩壊時のセルを黄色で塗る
                collapse_row_excel = collapse_index + 6 # Excelの行番号
                cell = worksheet.cell(row=collapse_row_excel, column=col_start+2)
                cell.fill = PatternFill(fgColor='FFFF00', fill_type='solid')
                
            color_folder_idx += 1

        # Excelファイルを保存
        writer.close()
        print(f"結果は {excel_path} に保存されました")
    except Exception as e:
        print(f"Excelファイルの保存に失敗しました: {str(e)}")

    print("全ての処理が完了しました")

if __name__ == "__main__":
    # 処理の設定 - バックスラッシュを使用
    base_path = r'C:\Research\exp_data\20250819'  # raw文字列として指定
    start_folder = 2
    end_folder = 116
    calibration = 39.5
    time_interval = 0.000005  # 時間間隔s（秒）
    start_image_num = 7  # t*=0とする画像番号

    blur_x = blur_y = 41
    binary_s = 31

    try:
        process_all_folders(start_folder, end_folder, base_path, calibration, time_interval, start_image_num,
                             excel_file_name="16_analysis_result.xlsx", blur_x = blur_x, blur_y = blur_y, binary_s = binary_s)
    except Exception as e:
        print(f"プログラム実行中にエラーが発生しました: {str(e)}")

    #calibration一覧
    #20231210(0.7) - 38
    #20231218(0.3) - 36
    #20250417(0.4) - 32.2
    #20250611(0.5) - 39.4
    #20250819(0.5) - 39.5