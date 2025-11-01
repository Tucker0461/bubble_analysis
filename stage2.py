import cv2
import numpy as np
import math
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import sys 

# T*計算関数 (変更なし)
def calculate_t_star(time, rmax):
    """
    t*を計算する関数 (フォルダごとのRmaxを使用)
    """
    rho = 1000 # 水の密度 (kg/m³)
    delta_p = 1e5 # 圧力差 (Pa)
    
    if rmax == 0:
        return 0
        
    denominator = 0.91468 * (rmax / 1000) * (rho / delta_p)**0.5
    return time / denominator

# 最大・崩壊点検出関数 (変更なし)
def find_bubble_points(radius_data):
    """
    気泡の最大半径点と最初の極小点（崩壊点）を見つける
    検索は画像の8番目（インデックス7）から開始し、最初の崩壊点までを有効範囲とする
    """
    max_radius = 0
    max_index = -1
    collapse_index = -1
    
    # Rmax/崩壊点の検索開始インデックスを 7 (8枚目) に設定
    START_INDEX = 10
    
    if len(radius_data) <= START_INDEX:
        return -1, -1 # 検索範囲がない
        
    # ----------------------------------------------
    # 1. 崩壊点の先行検出 (検索範囲の決定)
    # ----------------------------------------------
    # 崩壊点が見つかった場合、そこを検索の終了点とする
    end_search_index = len(radius_data)
    radius_threshold = 0.5

    for i in range(START_INDEX, len(radius_data)):
        
        is_local_minimum = False
        if i > START_INDEX and i < len(radius_data) - 1:
            if radius_data[i] < radius_threshold and radius_data[i] < radius_data[i-1] and radius_data[i] <= radius_data[i+1]:
                is_local_minimum = True
        elif i == len(radius_data) - 1 and radius_data[i] < radius_data[i-1] and i >= START_INDEX:
            is_local_minimum = True
            
        if is_local_minimum:
            collapse_index = i
            end_search_index = i + 1 # 崩壊点を含めて検索を終了
            break
            
        if radius_data[i] == 0 and i >= START_INDEX:
            if collapse_index == -1: 
                collapse_index = i
            end_search_index = i + 1
            break
            
    # ----------------------------------------------
    # 2. 決定された範囲でRmaxを検出
    # ----------------------------------------------
    
    temp_max_radius_mm = 0
    
    for i in range(START_INDEX, end_search_index):
        if i < len(radius_data):
            temp_max_radius_mm = max(temp_max_radius_mm, radius_data[i])
            
    # Rmax/崩壊点のインデックスを検出 (インデックス7/8枚目から検索)
    max_index = -1
    max_radius = 0
    started = False
    
    for i in range(START_INDEX, end_search_index):
        radius = radius_data[i]
        
        if not started and radius > 0:
            started = True
        if started:
            if radius > max_radius:
                max_radius = radius
                max_index = i
    
    return max_index, collapse_index

# --- 新規追加: 重み付けのための円のスライス面積を計算するヘルパー関数 ---
def _calculate_slice_area_weight(n, i):
    """
    直径 n の円を n 分割したときの上から i 番目 (i=1..n) のスライス面積 (重み) を計算する。
    一般的な弓形面積の公式に基づき、体積 V [pixel^3] の重みとする。
    """
    R = n / 2.0
    
    if R == 0:
        return 0.0

    # 1. 弓形面積 A(d) を計算するヘルパー (中心から距離 d で分割された小さい方の面積)
    def area_of_circular_segment(R, d_abs):
        if d_abs >= R:
            return 0.0
        
        # math.acos の引数 d_abs / R が [-1, 1] の範囲内であることを保証
        cos_theta_half = min(1.0, max(-1.0, d_abs / R))
        
        # 中心角の半分
        theta_half = math.acos(cos_theta_half)
        
        # 弓形面積 = R^2 * theta_half - d_abs * sqrt(R^2 - d_abs^2)
        return R**2 * theta_half - d_abs * math.sqrt(R**2 - d_abs**2)

    # 2. 上端から境界線までの累積面積 S(d) を計算
    # 境界線 d_signed の中心からの符号付き距離 (上側が正のR, 下側が負のR)
    # 境界線 i の位置 (y座標) は i (上から i 番目のピクセルの下側の境界)
    def cumulative_area_S(d_signed, R):
        d_abs = abs(d_signed)
        total_area = math.pi * R**2
        
        if d_abs >= R:
            return 0.0 if d_signed >= 0 else total_area 
            
        A_d = area_of_circular_segment(R, d_abs)
        
        if d_signed >= 0: # 中心より上: 小さい方の弓形面積
            return A_d
        else: # 中心より下: 大きい方の弓形面積
            return total_area - A_d

    # 3. スライス面積 ΔS_i = S(i) - S(i-1)
    # 線 i の中心からの符号付き距離 d_i = R - i
    # 線 i-1 の中心からの符号付き距離 d_{i-1} = R - (i-1) = d_i + 1

    d_i = R - i
    d_i_minus_1 = d_i + 1
    
    S_i = cumulative_area_S(d_i, R)
    S_i_minus_1 = cumulative_area_S(d_i_minus_1, R)
    
    return S_i - S_i_minus_1
# ------------------------------------------------------------------------


# 体積・重心計算関数 (変更済み: 重み付き算術平均)
def calculate_properties_from_binary(binary_path, wall_x_pixel, calibration, min_area_pixel2, max_individual_bubbles=4):
    
    empty_agg = {k: 0 for k in ['volume', 'radius', 'center_x', 'center_y', 'x_min_mm', 'x_max_mm', 'y_min_mm', 'y_max_mm', 'v_agarose', 'v_water', 'x_agarose', 'y_agarose', 'x_water', 'y_water', 'aspect_ratio']}

    if binary_path is None or not os.path.isfile(binary_path):
        return empty_agg, []
    
    try:
        binary = cv2.imdecode(np.fromfile(binary_path, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
        if binary is None:
            return empty_agg, []
            
        wall_x = int(wall_x_pixel) if wall_x_pixel is not None else -1 
        
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        all_bubble_data = []
        
        for contour in contours:
            area = cv2.contourArea(contour)
            if area >= min_area_pixel2:
                
                x, y, w, h = cv2.boundingRect(contour)
                mask = np.zeros(binary.shape, dtype=np.uint8)
                cv2.drawContours(mask, [contour], 0, 255, -1)
                
                # --- 新しい計算ロジックのための集計変数 ---
                total_weight_V = 0.0 
                moment_x_total = 0.0 
                moment_y_total = 0.0 
                
                # ゾーン別集計
                v_agarose_sum_weight = 0.0
                m_x_agarose = 0.0
                m_y_agarose = 0.0
                v_water_sum_weight = 0.0
                m_x_water = 0.0
                m_y_water = 0.0
                # ------------------------------------------

                # ピクセル列 (x) ごとの処理
                for x_abs in range(x, x + w):
                    col_data = mask[y:y+h, x_abs] 
                    
                    if np.sum(col_data) > 0:
                        
                        # 連続する黒ピクセルの塊 (ラン) を検出する
                        current_run_start_rel = -1
                        run_list = [] # [(start_y_rel, run_length_n), ...]
                        
                        for y_rel in range(h):
                            is_black = col_data[y_rel] > 0
                            
                            if is_black:
                                if current_run_start_rel == -1:
                                    current_run_start_rel = y_rel
                            else: # is_white
                                if current_run_start_rel != -1:
                                    # 塊の終わり
                                    run_length_n = y_rel - current_run_start_rel
                                    run_list.append((current_run_start_rel, run_length_n))
                                    current_run_start_rel = -1

                        # 列の端で塊が終わる場合の処理
                        if current_run_start_rel != -1:
                            run_length_n = h - current_run_start_rel
                            run_list.append((current_run_start_rel, run_length_n))

                        # 塊 (ラン) ごとの重み計算
                        for start_y_rel, n in run_list:
                            if n == 0: continue
                            
                            # 塊内の i=1 から n までのピクセルを処理
                            for i in range(1, n + 1):
                                # i: 塊内での上からのスライス番号 (1-based)
                                # y_rel: 画像全体での相対 y 座標 (0-based)
                                y_rel = start_y_rel + i - 1
                                y_abs = y + y_rel # 画像全体での絶対 y 座標
                                
                                # 重み wij = ΔS_i(n) を計算
                                weight_wij = _calculate_slice_area_weight(n, i)
                                
                                # --- 全体集計 ---
                                total_weight_V += weight_wij
                                moment_x_total += weight_wij * x_abs
                                moment_y_total += weight_wij * y_abs
                                
                                # --- ゾーン別集計 ---
                                if wall_x != -1:
                                    if x_abs < wall_x: # Agarose
                                        v_agarose_sum_weight += weight_wij
                                        m_x_agarose += weight_wij * x_abs
                                        m_y_agarose += weight_wij * y_abs
                                    elif x_abs >= wall_x: # Water
                                        v_water_sum_weight += weight_wij
                                        m_x_water += weight_wij * x_abs
                                        m_y_water += weight_wij * y_abs
                                else: # 壁情報がない場合はすべて Agarose 側に集計
                                    v_agarose_sum_weight += weight_wij
                                    m_x_agarose += weight_wij * x_abs
                                    m_y_agarose += weight_wij * y_abs


                # ----------------------------------------------------
                # 2. 集計データの計算 (Contourごと)
                # ----------------------------------------------------
                
                if total_weight_V > 0:
                    center_x_pixel = moment_x_total / total_weight_V
                    center_y_pixel = moment_y_total / total_weight_V
                    
                    aspect_ratio = h / w if w != 0 else 0
                else:
                    center_x_pixel, center_y_pixel, aspect_ratio = 0, 0, 0

                all_bubble_data.append({
                    'volume_pixel3': total_weight_V, 
                    'center_x_pixel': center_x_pixel,
                    'center_y_pixel': center_y_pixel,
                    'aspect_ratio': aspect_ratio,
                    'min_x_pix': x, 'max_x_pix': x + w, 
                    'min_y_pix': y, 'max_y_pix': y + h,
                    'v_agarose_pix3': v_agarose_sum_weight,
                    'm_x_agarose': m_x_agarose,
                    'm_y_agarose': m_y_agarose,
                    'v_water_pix3': v_water_sum_weight,
                    'm_x_water': m_x_water,
                    'm_y_water': m_y_water,
                })

        # ----------------------------------------------------
        # 3. フレーム集計データ (シート1用)
        # ----------------------------------------------------
        
        total_volume_pixel3 = sum(d['volume_pixel3'] for d in all_bubble_data)
        
        # 重み付き重心 (フレーム全体)
        total_moment_x = sum(d['volume_pixel3'] * d['center_x_pixel'] for d in all_bubble_data)
        total_moment_y = sum(d['volume_pixel3'] * d['center_y_pixel'] for d in all_bubble_data)
        
        v_agarose_sum_pix3 = sum(d['v_agarose_pix3'] for d in all_bubble_data)
        m_x_agarose_sum = sum(d['m_x_agarose'] for d in all_bubble_data)
        m_y_agarose_sum = sum(d['m_y_agarose'] for d in all_bubble_data)
        
        v_water_sum_pix3 = sum(d['v_water_pix3'] for d in all_bubble_data)
        m_x_water_sum = sum(d['m_x_water'] for d in all_bubble_data)
        m_y_water_sum = sum(d['m_y_water'] for d in all_bubble_data)
        
        global_min_x = min(d['min_x_pix'] for d in all_bubble_data) if all_bubble_data else 0
        global_max_x = max(d['max_x_pix'] for d in all_bubble_data) if all_bubble_data else 0
        global_min_y = min(d['min_y_pix'] for d in all_bubble_data) if all_bubble_data else 0
        global_max_y = max(d['max_y_pix'] for d in all_bubble_data) if all_bubble_data else 0
        
        if total_volume_pixel3 > 0:
            volume_total_mm3 = total_volume_pixel3 / (calibration**3)
            
            # 全体の重心 (ピクセル単位の重み付き平均)
            center_x_pixel_agg = total_moment_x / total_volume_pixel3
            center_y_pixel_agg = total_moment_y / total_volume_pixel3
            
            center_x_mm = center_x_pixel_agg / calibration
            center_y_mm = center_y_pixel_agg / calibration
            
            equivalent_radius_mm = (3 * volume_total_mm3 / (4 * math.pi))**(1/3) 
            
            v_agarose_mm3 = v_agarose_sum_pix3 / (calibration**3)
            v_water_mm3 = v_water_sum_pix3 / (calibration**3)
            
            # ゾーン別重心
            x_agarose_mm = (m_x_agarose_sum / v_agarose_sum_pix3) / calibration if v_agarose_sum_pix3 > 0 else 0
            y_agarose_mm = (m_y_agarose_sum / v_agarose_sum_pix3) / calibration if v_agarose_sum_pix3 > 0 else 0
            
            x_water_mm = (m_x_water_sum / v_water_sum_pix3) / calibration if v_water_sum_pix3 > 0 else 0
            y_water_mm = (m_y_water_sum / v_water_sum_pix3) / calibration if v_water_sum_pix3 > 0 else 0
            
            # 最大気泡のアスペクト比
            max_bubble = max(all_bubble_data, key=lambda x: x['volume_pixel3'])
            max_bubble_ar = max_bubble['aspect_ratio']
            
        else:
            volume_total_mm3 = equivalent_radius_mm = center_x_mm = center_y_mm = 0
            v_agarose_mm3 = v_water_mm3 = x_agarose_mm = y_agarose_mm = x_water_mm = y_water_mm = max_bubble_ar = 0
            
        aggregate_data = {
            'volume': volume_total_mm3,
            'radius': equivalent_radius_mm,
            'center_x': center_x_mm,
            'center_y': center_y_mm,
            'x_min_mm': global_min_x / calibration,
            'x_max_mm': global_max_x / calibration,
            'y_min_mm': global_min_y / calibration,
            'y_max_mm': global_max_y / calibration,
            'v_agarose': v_agarose_mm3,
            'v_water': v_water_mm3,
            'x_agarose': x_agarose_mm,
            'y_agarose': y_agarose_mm,
            'x_water': x_water_mm,
            'y_water': y_water_mm,
            'aspect_ratio': max_bubble_ar,
        }

        # ----------------------------------------------------
        # 4. 個別気泡データ (シート2用)
        # ----------------------------------------------------
        
        all_bubble_data.sort(key=lambda x: x['volume_pixel3'], reverse=True)
        
        individual_bubble_data = []
        
        for d in all_bubble_data[:max_individual_bubbles]:
            individual_bubble_data.append({
                'volume': d['volume_pixel3'] / (calibration**3),
                'center_x': d['center_x_pixel'] / calibration,
                'center_y': d['center_y_pixel'] / calibration,
            })
            
        return aggregate_data, individual_bubble_data
    
    except Exception as e:
        print(f"体積計算でエラーが発生しました: {binary_path} - {str(e)}")
        return empty_agg, []

# ==============================================================================
# フォルダごとのRmaxを使用するよう修正されたメイン関数 (変更なし)
# ==============================================================================

def stage2_main(base_path, start_folder, end_folder, calibration, time_interval, start_image_num, min_area_pixel2, max_individual_bubbles=4, excel_file_name='analysis_result.xlsx'):
    
    # ... (stage2_main関数の既存ロジックは変更なし) ...
    # 省略... (main関数ブロックは元のコードと同一)

    # 基準値の読み込み (変更なし)
    reference_path = os.path.join(base_path, 'reference.xlsx').replace('/', '\\')
    if not os.path.isfile(reference_path):
        print(f"基準値ファイルが存在しません: {reference_path}")
        return
        
    try:
        reference_df = pd.read_excel(reference_path)
        reference_dict = dict(zip(reference_df['Folder'], reference_df['base_x(pix)']))
        print("Stage 2: 基準値の読み込みが完了しました")
    except Exception as e:
        print(f"Stage 2: 基準値ファイルの読み込みに失敗しました: {str(e)}")
        return

    results_base_path = os.path.join(base_path, "results").replace('/', '\\')
    binary_base_path = os.path.join(base_path, "binary_images").replace('/', '\\')
    os.makedirs(results_base_path, exist_ok=True)
    
    all_aggregate_data = {}
    all_individual_data = {}
    max_radii = {} # 各フォルダの最大半径 Rmax (t*計算用)
    gammas = {}
    
    excel_path = os.path.join(results_base_path, excel_file_name).replace('/', '\\')

    print("\n--- Stage 2: 体積・重心の計算とExcelへの出力を開始 ---")
    
    folder_list = list(range(start_folder, end_folder + 1))
    
    # ----------------------------------------------------
    # A. t*計算のための Rmax 決定（フォルダごと）(変更なし)
    # ----------------------------------------------------
    
    for folder_num in folder_list:
        folder_name = str(folder_num)
        binary_folder = os.path.join(binary_base_path, folder_name).replace('/', '\\')
        
        if not os.path.exists(binary_folder): 
            max_radii[folder_name] = 0
            continue
        
        current_wall_x_pixel = reference_dict.get(folder_num, None)
        file_list = sorted([f for f in os.listdir(binary_folder) if f.endswith('_binary.bmp')])
        
        # 1. 全フレームの半径データを取得
        radius_data_all = []
        for file_name in file_list:
            binary_path = os.path.join(binary_folder, file_name).replace('/', '\\')
            # calculate_properties_from_binary は新しいロジックに置き換わっているが、radiusの計算は継承
            agg_data, _ = calculate_properties_from_binary(binary_path, current_wall_x_pixel, calibration, min_area_pixel2, 0)
            radius_data_all.append(agg_data.get('radius', 0) if agg_data else 0)
        
        # 2. 8枚目以降、崩壊点までの範囲で最大半径 Rmax を決定
        temp_max_radius_mm = 0
        
        # Rmax/崩壊点のインデックスを検出 (インデックス7/8枚目から検索)
        max_index, collapse_index = find_bubble_points(radius_data_all)
        
        # 検索開始インデックス: 10 (11枚目)
        START_INDEX = 10

        # 検索終了インデックス: 崩壊点が見つかった場合、そのインデックス
        end_index = collapse_index if collapse_index != -1 else len(radius_data_all)
        
        # 11枚目から崩壊点までの範囲で最大半径を計算
        for i in range(START_INDEX, end_index):
            if i < len(radius_data_all):
                temp_max_radius_mm = max(temp_max_radius_mm, radius_data_all[i])
            
        # フォルダごとの Rmax を保存
        max_radii[folder_name] = temp_max_radius_mm

    # ----------------------------------------------------
    # B. 最終データ収集のためのループ (変更なし)
    # ----------------------------------------------------
    
    for folder_num in folder_list:
        folder_name = str(folder_num)
        binary_folder = os.path.join(binary_base_path, folder_name).replace('/', '\\')
        
        current_rmax_for_t_star = max_radii.get(folder_name, 0) # フォルダRmaxを取得
        
        # フォルダの最大半径 Rmax が 0 の場合、またはフォルダが存在しない場合はスキップ
        if current_rmax_for_t_star == 0 or not os.path.exists(binary_folder): continue

        print(f"フォルダ {folder_name} のデータ処理中...")
        
        current_wall_x_pixel = reference_dict.get(folder_num, None)
        file_list = sorted([f for f in os.listdir(binary_folder) if f.endswith('_binary.bmp')])
        
        folder_agg_data = []
        folder_indiv_data = []
        
        for idx, file_name in enumerate(file_list):
            
            # --- 修正ロジックの適用: 7枚目までの集計値は0 ---
            if idx < start_image_num - 1: # start_image_num=7 の場合、idx=0から5 (1枚目から6枚目)
                # 7枚目まではファイルパスをNoneとし、calculate_properties_from_binaryで全0データを得る
                agg_data, indiv_data = calculate_properties_from_binary(None, None, calibration, min_area_pixel2, 0)
            
            elif idx == start_image_num - 1: # idx=6 (7枚目)
                # 7枚目のみ、集計データは0にするが、個別データも0に
                agg_data, indiv_data = calculate_properties_from_binary(None, None, calibration, min_area_pixel2, 0)

            else: # idx >= start_image_num (8枚目以降)
                # 8枚目以降は通常通り計算
                binary_path = os.path.join(binary_folder, file_name).replace('/', '\\')
                agg_data, indiv_data = calculate_properties_from_binary(binary_path, current_wall_x_pixel, calibration, min_area_pixel2, max_individual_bubbles)
            # --------------------------

            # t* の計算
            if idx < start_image_num - 1: # 7枚目 (インデックス6) まで
                t_star = 0
            else: # 8枚目 (インデックス7) 以降
                elapsed_time = (idx - (start_image_num - 1)) * time_interval
                t_star = calculate_t_star(elapsed_time, current_rmax_for_t_star) 
            
            if agg_data:
                agg_data['t_star'] = t_star 
                folder_agg_data.append(agg_data) 
            
            # 個別気泡データ
            folder_indiv_data.append({
                't_star': t_star,
                'bubbles': indiv_data
            })
        
        all_aggregate_data[folder_name] = folder_agg_data
        all_individual_data[folder_name] = folder_indiv_data
        
        # ------------------------------------------------------------------------
        # γの計算 (変更なし)
        # ------------------------------------------------------------------------
        gamma = 0
        current_rmax = current_rmax_for_t_star # Rmax (分母)
        INITIAL_X_INDEX = 9 # 10枚目

        if len(folder_agg_data) > INITIAL_X_INDEX and folder_num in reference_dict and current_rmax > 0:
            
            base_x_mm = reference_dict[folder_num] / calibration
            
            # 10枚目 (インデックス9) の全体の集計データ (aggregate_data) を取得
            initial_frame_agg_data = folder_agg_data[INITIAL_X_INDEX] 
            
            # 全体の重心X座標 ('center_x') を取得して使用
            initial_x = initial_frame_agg_data.get('center_x', 0)
            
            # initial_x が 0 でないこと（気泡が検出されていること）を確認してから計算
            if initial_x != 0:
                gamma = (initial_x - base_x_mm) / current_rmax 

            else:
                print(f"警告: フォルダ {folder_num} の10枚目で全体の重心X座標 (center_x) が 0 のため、γを 0 としました。")
            
            gammas[folder_name] = gamma
        else:
            gammas[folder_name] = 0
        # ------------------------------------------------------------------------

    # ----------------------------------------------------
    # C. Excel出力処理 (変更なし)
    # ----------------------------------------------------
    
    agg_cols_data = ['t_star', 'volume', 'radius', 'center_x', 'center_y', 'x_min_mm', 'x_max_mm', 'y_min_mm', 'y_max_mm', 'v_agarose', 'v_water', 'x_agarose', 'y_agarose', 'x_water', 'y_water', 'aspect_ratio']
    num_indiv_cols_per_bubble = 3 
    VOLUME_DATA_INDEX = agg_cols_data.index('volume') 
    
    RED_FILL = PatternFill(fgColor='FF0000', fill_type='solid')
    YELLOW_FILL = PatternFill(fgColor='FFFF00', fill_type='solid')
    
    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1_Aggregate'
        ws2 = wb.create_sheet(title='Sheet2_Individual')
        
        current_col = 1
        current_col_s2 = 1
        
        for folder_num in folder_list:
            folder_name = str(folder_num)
            if folder_name not in all_aggregate_data: 
                continue 
                
            agg_data = all_aggregate_data[folder_name]
            indiv_data = all_individual_data[folder_name]

            # ---------------------
            # Sheet 1: データ書き込みと色付け
            # ---------------------
            
            # 1. メタデータ (行1-4)
            ws1.cell(row=1, column=current_col + 1, value='Folder')
            ws1.cell(row=2, column=current_col + 1, value='Rmax')
            ws1.cell(row=3, column=current_col + 1, value='γ')
            
            ws1.cell(row=1, column=current_col + 2, value=folder_num)
            ws1.cell(row=2, column=current_col + 2, value=max_radii.get(folder_name, 0))
            ws1.cell(row=3, column=current_col + 2, value=gammas.get(folder_name, 0))
            
            # 2. ヘッダー (行5)
            for col_idx, col_name in enumerate(agg_cols_data):
                ws1.cell(row=5, column=current_col + col_idx, value=col_name)

            # 3. データ本体 (行6以降)
            radius_data = [d.get('radius', 0) for d in agg_data]
            max_index, collapse_index = find_bubble_points(radius_data) 
            
            VOLUME_DATA_INDEX_IN_LOOP = agg_cols_data.index('volume')
            
            for row_idx in range(120):
                data_row = row_idx + 6 # Excelの行番号
                
                if row_idx < len(agg_data):
                    frame_data = agg_data[row_idx]
                    
                    # 4. データ書き込み
                    for col_idx, col_name in enumerate(agg_cols_data):
                        value = frame_data.get(col_name)
                        cell = ws1.cell(row=data_row, column=current_col + col_idx, value=value)
                        
                        # --- 色付けロジック ---
                        if col_idx == VOLUME_DATA_INDEX_IN_LOOP:
                            if row_idx == max_index:
                                cell.fill = RED_FILL
                            elif row_idx == collapse_index:
                                cell.fill = YELLOW_FILL
            
            # 次のフォルダの開始列へ
            current_col += len(agg_cols_data)

            # ---------------------
            # Sheet 2: データ書き込み
            # ---------------------
            
            # 1. 1行目にデータ番号 (Folder) を挿入
            ws2.cell(row=1, column=current_col_s2).value = folder_num

            # 2. ヘッダー (行4, 5)
            ws2.cell(row=4, column=current_col_s2, value='t_star') # t*列

            for b_idx in range(max_individual_bubbles):
                col_offset = b_idx * num_indiv_cols_per_bubble
                
                # 4行目: Bubble ID (t*列の次)
                ws2.cell(row=4, column=current_col_s2 + 1 + col_offset, value=f'Bubble {b_idx+1}')
                
                # 5行目: Data Items
                for item_idx, item_name in enumerate(['volume', 'center_x', 'center_y']):
                    ws2.cell(row=5, column=current_col_s2 + 1 + col_offset + item_idx, value=item_name)
            
            # 3. データ本体 (行6以降)
            for row_idx in range(120):
                data_row = row_idx + 6 
                
                if row_idx < len(indiv_data):
                    frame_data = indiv_data[row_idx]
                    
                    # t* 列 (current_col_s2)
                    ws2.cell(row=data_row, column=current_col_s2, value=frame_data['t_star'])
                    
                    # 個別気泡データ
                    for b_idx in range(max_individual_bubbles):
                        if b_idx < len(frame_data['bubbles']):
                            bubble = frame_data['bubbles'][b_idx]
                            col_offset = b_idx * num_indiv_cols_per_bubble
                            
                            ws2.cell(row=data_row, column=current_col_s2 + 1 + col_offset, value=bubble['volume'])
                            ws2.cell(row=data_row, column=current_col_s2 + 1 + col_offset + 1, value=bubble['center_x'])
                            ws2.cell(row=data_row, column=current_col_s2 + 1 + col_offset + 2, value=bubble['center_y'])

            # Sheet 2の次のフォルダ列へ (t*列 + Bubbleデータ列)
            current_col_s2 += 1 + max_individual_bubbles * num_indiv_cols_per_bubble

        # Workbookを保存
        wb.save(excel_path)
        
        print(f"結果は {excel_path} に保存されました。")
    except Exception as e:
        print(f"致命的なエラー: Excelファイルの保存に失敗しました。パスを確認してください: {excel_path}")
        print(f"エラー詳細: {str(e)}")
        sys.exit(1)

    print("Stage 2 完了: 全ての処理が完了しました")

if __name__ == "__main__":
    # --- Stage 2 設定 ---
    base_path = r'C:\Research\exp_data\20250611' 
    start_folder = 2
    end_folder = 150

    # 計算パラメータ
    calibration = 39.4
    time_interval = 0.000005
    start_image_num = 7 # t*=0とする画像番号 (1枚目から7枚目までが集計値0、8枚目以降で計算)
    min_area_pixel2 = 0
    max_individual_bubbles = 4

    try:
        stage2_main(base_path, start_folder, end_folder, calibration, time_interval, start_image_num,
                    min_area_pixel2, max_individual_bubbles, excel_file_name="6_analysis_20250611.xlsx")
    except Exception as e:
        print(f"プログラム実行中にエラーが発生しました: {str(e)}")

    #calibration一覧
    #20231210(0.7) - 38
    #20250417(0.4) - 32.2
    #20250611(0.5) - 39.4
    #20250819(0.5) - 39.5